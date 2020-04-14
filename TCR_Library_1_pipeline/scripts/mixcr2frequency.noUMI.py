import os
import sys
import openpyxl
import math
from itertools import islice


def usage():
    """docstring for usage

    python mixcr2frequency.noUMI.py G67E1L1.mixcr.out.clonotypes.TRA.count.txt > G67E1L1.filtered2reads.freq.xlsx

    20200109: added scater chart.
    20200108: created
    """


def addtwodimdict(thedict, key_a, key_b, val):
    ''' this is a function to add two dimetion dict '''
    if key_a in thedict:
        thedict[key_a].update({key_b: val})
    else:
        thedict.update({key_a: {key_b: val}})
    return thedict


def dict_freqency(adict):
    total_sum = sum((float(k) for k in adict.values()))
    new_dict = {}
    for each in adict:
        new_dict[each] = float(adict[each])/total_sum
    return new_dict


def deal_file(count_file):
    """docstring for deal_file"""
    a_dict = {}
    with open(count_file, "r") as f:
        for line in islice(f, 1, None):  # skip header line.
            line = line.rstrip("\n")
            Clonotype, TRV, CDR3, TRJ, ReadsNumber = line.split("\t")
            a_dict[Clonotype] = ReadsNumber

    transformated_dict = dict_freqency(a_dict)
    '''
    print("Clonotype\tSupporting_Reads\tReads_Frequency\tFrequency_Log10")

    for k in transformated_dict:
        print("{}\t{}\t{}\t{}".format(k, a_dict[k], transformated_dict[k], math.log(transformated_dict[k], 10)))
    '''
    return a_dict, transformated_dict


def process_worksheet(sheet, a_dict, t_dict):
    """docstring for process_worksheet"""
    title = ["Clonotype", "Supporting_Reads", "Reads_Frequency", "Frequency_Log10"]
    for i, each in enumerate(title, start=1):
        sheet.cell(row=1, column=i, value=each)

    next_row = 2
    for each in sorted(t_dict.items(), key=lambda item:item[1], reverse=True):
        sheet.cell(column=1, row=next_row, value=each[0])
        sheet.cell(column=2, row=next_row, value=a_dict[each[0]])
        sheet.cell(column=3, row=next_row, value=each[1])
        sheet.cell(column=4, row=next_row, value=math.log(each[1], 10))
        next_row += 1


def create_charts(sheet, chart_type):
    """docstring for create_charts"""
    if chart_type == 'Scatter':
        chart = openpyxl.chart.ScatterChart()
    elif chart_type == 'Line':
        chart = openpyxl.chart.LineChart()
    elif chart_type == 'Bar':
        chart = openpyxl.chart.Bar()

    values = openpyxl.chart.Reference(sheet, min_col=4, min_row=2, max_col=4, max_row=sheet.max_row)
    chart.hight = 30
    chart.legend = None
    chart.add_data(values, titles_from_data=False)
    chart.title = "Frequency_Log10"
    chart.style = 11

    sheet.add_chart(chart, "F6")


def main():
    """docstring for main"""

    count_file = sys.argv[1]
    file_name = os.path.basename(count_file).split(".")[0]
    a_dict, transformated_dict = deal_file(count_file)

    # write to excel
    work_book = openpyxl.Workbook()
    work_book.create_sheet(index=0, title="sheet1")
    sheet = work_book.active
    process_worksheet(sheet, a_dict, transformated_dict)
    
    # chart_type = 'Scatter'
    chart_type = 'Line'
    create_charts(sheet, chart_type)
    work_book.save(file_name + '.filtered2reads.freq.xlsx')


if __name__ == '__main__':
    main()

