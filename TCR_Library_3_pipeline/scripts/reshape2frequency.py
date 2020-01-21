import os
import sys
from itertools import islice
import math
import openpyxl


def usage():
    """
    docstring for usage:

    python  reshape2frequency.py G41E1L2.merged.umi.count.reshape.xls 

    20200108: add draw line chart in each sheet.
    20191215: add openpyxl module. creat excel file automatically.
    20190805: create this file.
    """


def dict_freqency(adict):
    total_sum = sum((int(k) for k in adict.values()))
    new_dict = {}
    for each in adict:
        new_dict[each] = float(adict[each])/total_sum
    return new_dict


def process_worksheet(sheet, adict, tdict, mdict):
    """docstring for process_worksheet"""
    # write the header.
    title = ["Clonotype","Supporting_Reads","Reads_Frequency","Frequency_Log10","Molecules"]
    for i, each in enumerate(title, start=1):
        sheet.cell(row=1,column=i).value = each

    # write the value.
    row_index = 2
    for each in sorted(tdict.items(), key=lambda item:item[1], reverse=True):
        sheet.cell(row=row_index, column=1, value=each[0])
        sheet.cell(row=row_index, column=2, value=adict[each[0]])
        sheet.cell(row=row_index, column=3, value=each[1])
        sheet.cell(row=row_index, column=4, value=mdict[each[0]])
        sheet.cell(row=row_index, column=5, value=math.log(each[1],10))
        row_index += 1


def create_charts(sheet):
    """docstring for create_charts"""
    c1 = openpyxl.chart.LineChart()
    c1.title = "Frequency_Log10"
    c1.style = 11
    c1.hight=30

    data = openpyxl.chart.Reference(sheet, min_col=sheet.max_column, min_row=1, max_row=sheet.max_row)
    c1.add_data(data, titles_from_data=False)
    c1.legend = None
    sheet.add_chart(c1, "G6")


def deal_input_file(afile):
    adict = {}
    tdict = {}
    mdict = {}
    with open(afile, "r") as f:
        for line in islice(f, 1, None):
            line = line.rstrip("\n")
            Clonotype, TRV, CDR3, TRJ, UMIcount, ReadsNumber = line.split("\t")
            adict[Clonotype] = ReadsNumber
            mdict[Clonotype] = UMIcount
    tdict = dict_freqency(adict)
    return adict, tdict, mdict

def main():
    """docstring for main"""
    # deal with data.
    input_file = sys.argv[1]  # G39E1L2.merged.umi.count.reshape.xls
    file_name = os.path.basename(input_file).split(".")[0]
    adict,tdict,mdict = deal_input_file(input_file)

    # write to excel.
    work_book = openpyxl.Workbook()
    work_book.create_sheet(index=0, title="sheet1")
    sheet = work_book.active
    process_worksheet(sheet, adict, tdict, mdict)

    # plot a charts.
    create_charts(sheet)
    work_book.save(file_name+ '.filtered2reads.freq.xlsx')


if __name__ == '__main__':
    main()
