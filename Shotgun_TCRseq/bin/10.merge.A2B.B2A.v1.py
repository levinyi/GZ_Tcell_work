import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl.styles import numbers


def usage():
    print('''usage:
    python {0} <file1> <file2> <output file xlsx>
example:
    python {0} FromA2B/Total.pairs.FromA2B.threshold.50.add.filtered.sharedwells.3.xls \\
    FromB2A/Total.pairs.FromB2A.threshold.50.add.filtered.sharedwells.3.xls  \\
    A2B_B2A_detail.threshold.50.sharedwells.3.new.xlsx

Updates:
    20200610    optimized code. fix some bugs.
        '''.format(os.path.basename(sys.argv[0])))


wb = Workbook()
ws = wb.active

a2b_file = sys.argv[1]  #
b2a_file = sys.argv[2]
outfilename = sys.argv[3]

adict = {}
alist = []
with open(a2b_file, "r") as f1:
    for line in f1:
        line = line.rstrip("\n")
        colum_1 = line.split()[0]
        colum_3 = line.split()[3]
        keya = colum_1 + '_' + colum_3
        alist.append(keya)
        adict[keya] = line

bdict = {}
with open(b2a_file, "r") as f2:
    for line in f2:
        line = line.rstrip("\n")
        colum_1 = line.split()[0]
        colum_3 = line.split()[3]
        keyb = colum_3 + '_' + colum_1
        alist.append(keyb)
        bdict[keyb] = line

alist = list(set(alist))
style_both = Font(color=colors.BLUE, bold=True)
style_red = Font(color=colors.RED)
style_green = Font(color=colors.GREEN)
title = ["Note","TRA_clone","TRA_count","TRA_clone_wells","TRB_clone","TRB_count","TRB_clone_wells","shared_wells","p1","ratio","p2","Cell.Num.Freq","Theoretical.Well.Number","TRA-dropout","TRB-dropout","Theoretical.Cell.Num.Freq"]

for col in range(1, len(title)+1):
    ws.cell(row=1, column=col,).value = title[col-1]
row = 2
col = 1
for each in alist:
    if each in adict and each in bdict:
        # print("both\t{}".format(adict[each]))
        ws.cell(row=row, column=1).value = "both"
        ws.cell(row=row, column=1).font = style_both
        alist = []
        for x in adict[each].split("\t"):
            ws.cell(row=row, column=col+1).value = x
            ws.cell(row=row, column=col+1).number_format = "#,###"
            ws.cell(row=row, column=col+1).font = style_both
            col += 1
    elif each in adict and each not in bdict:
        ws.cell(row=row, column=1).value = "onlyA2B"
        ws.cell(row=row, column=1).font = style_green
        for x in adict[each].split("\t"):
            ws.cell(row=row, column=col+1).value = x
            ws.cell(row=row, column=col+1).number_format= "#,###"
            ws.cell(row=row, column=col+1).font = style_green
            col += 1
        # print("onlyA2B\t{}".format(adict[each]))
    elif each not in adict and each in bdict:
        bclone,bcount,bwells,aclone,acount,awells,shared_wells,p1,ratio,p2,anum,bnum,bdropout,adropout,cfreq = bdict[each].split("\t")
        alist = [aclone,acount,awells,bclone,bcount,bwells,shared_wells,p1,ratio,p2,anum,bnum,adropout,bdropout,cfreq]
        ws.cell(row=row, column=1).value = "onlyB2A"
        ws.cell(row=row, column=1).font = style_red
        for x in alist:
            ws.cell(row=row, column=col+1).value = x
            ws.cell(row=row, column=col+1).number_format = "#,###"
            ws.cell(row=row, column=col+1).font = style_red
            col += 1
        '''
        print("onlyB2A\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}".format(
            aclone,acount,awells,bclone,bcount,bwells,shared_wells,p1,ratio,p2,anum,bnum,adropout,bdropout,cfreq))
        '''
    row += 1
    col = 1
wb.save(outfilename)
