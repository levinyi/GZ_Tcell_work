import sys
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl.styles import numbers

def usage():
    print('''Usage:
    python {0} <xlsx file> \\
            <TRA_freq_file1>\\
             <TRB_freq_file1> \\
             <TRA_freq_file2> \\
             <TRB_freq_file2> \\
             <output.xlsx>

Example:
    python {0} 

Updates:
    20200610    Created.
    '''.format(os.path.basename(sys.argv[0])))


def deal_freq_file(afile):
    adict = {}
    with open(afile, "r") as f:
        for line in f:
            line = line.rstrip("\n")
            clone, count, freq = line.split()
            adict[clone] = freq
    return adict

def main():
    if len(sys.argv) !=7:
        usage()
        sys.exit("Error: Wrong input file. please check: 7 input files but you put {0}".format(len(sys.argv)))
    
    freq1_file = sys.argv[2]
    freq2_file = sys.argv[3]
    freq3_file = sys.argv[4]
    freq4_file = sys.argv[5]
    freq1_dict = deal_freq_file(freq1_file)
    freq2_dict = deal_freq_file(freq2_file)
    freq3_dict = deal_freq_file(freq3_file)
    freq4_dict = deal_freq_file(freq4_file)

    wb = openpyxl.load_workbook(sys.argv[1])
    sheet = wb.active
    tra_column = 2
    trb_column = 5
    out1_column = sheet.max_column + 1
    out2_column = sheet.max_column + 2
    out3_column = sheet.max_column + 3
    out4_column = sheet.max_column + 4

# write title.
    title = ["TRA.frequency.in.G","TRB.frequency.in.G","TRA.frequency.in.G","TRB.frequency.in.G"]
    maxcolumn = sheet.max_column
    for col in range(1, len(title)+1):
        sheet.cell(row=1, column=maxcolumn + col,).value = title[col-1]

# write column value:
    for row in sheet.iter_rows(min_row=2):
        tra_value = sheet.cell(row=row[0].row, column=tra_column).value
        trb_value = sheet.cell(row=row[0].row, column=trb_column).value
        # print(tra_value,trb_value)
        v1 = freq1_dict.get(tra_value, 0)
        v2 = freq2_dict.get(trb_value, 0)
        v3 = freq3_dict.get(tra_value, 0)
        v4 = freq4_dict.get(trb_value, 0)
        sheet.cell(row=row[0].row, column=out1_column).value = v1
        sheet.cell(row=row[0].row, column=out2_column).value = v2
        sheet.cell(row=row[0].row, column=out3_column).value = v3
        sheet.cell(row=row[0].row, column=out4_column).value = v4
        
    wb.save(sys.argv[6])


if __name__ == '__main__':
    main()
